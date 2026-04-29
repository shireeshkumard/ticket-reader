from pathlib import Path

import openpyxl

from .models import Ticket


def read_xlsx(filepath: str | Path) -> list[Ticket]:
    """Read tickets from an Excel (.xlsx) file.

    Args:
        filepath: Path to the ``.xlsx`` file. Headers are expected on
            **row 2** (row 1 is typically a title or merged-cell row).
            Data rows start on row 3.
            Required column: ``Ticket number``. Optional columns:
            ``Ticket description``, ``Status``, ``Current Owner``, ``Issue type``.
            Column matching is case-insensitive and whitespace-tolerant.

    Returns:
        List of :class:`Ticket` objects. Blank rows and note rows without a
        ticket number are skipped automatically.

    Raises:
        FileNotFoundError: If the file does not exist.
        ValueError: If no ``Ticket number`` column is found.

    Example::

        from ticket_reader import read_xlsx

        for ticket in read_xlsx("Tickets.xlsx"):
            print(ticket.ticket_number, ticket.issue_type)
    """
    filepath = Path(filepath)
    if not filepath.exists():
        raise FileNotFoundError(f"No such file: '{filepath}'")

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    tickets: list[Ticket] = []

    try:
        headers = [cell.value for cell in ws[2]]
        cols: dict[str, int] = {
            str(h).strip().lower(): i
            for i, h in enumerate(headers)
            if h is not None
        }

        ticket_idx = cols.get("ticket number")
        if ticket_idx is None:
            raise ValueError(
                f"No 'Ticket number' column found in '{filepath}'. "
                f"Headers found: {[h for h in headers if h]}"
            )

        desc_idx = cols.get("ticket description")
        status_idx = cols.get("status")
        owner_idx = cols.get("current owner")
        type_idx = cols.get("issue type")

        def cell(row: tuple, idx: int | None) -> str:
            if idx is None or idx >= len(row):
                return ""
            val = row[idx]
            return str(val).strip() if val is not None else ""

        for row in ws.iter_rows(min_row=3, values_only=True):
            num = cell(row, ticket_idx)
            if not num or num.lower() == "ticket number":
                continue

            tickets.append(Ticket(
                ticket_number=num,
                description=cell(row, desc_idx),
                status=cell(row, status_idx),
                owner=cell(row, owner_idx),
                issue_type=cell(row, type_idx),
            ))
    finally:
        wb.close()

    return tickets
