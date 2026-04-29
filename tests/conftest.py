import pytest
import openpyxl
from pathlib import Path


@pytest.fixture
def sample_xlsx(tmp_path: Path) -> Path:
    """Minimal valid XLSX — headers on row 2, data from row 3."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Ticket Report"])
    ws.append(["Ticket number", "Ticket description", "Status", "Current Owner", "Issue type"])
    ws.append(["TSR-001", "Login fails on mobile", "Open", "PACS team", "Blocker"])
    ws.append(["TSR-002", "App crashes on startup", "Resolved", "Product", "Critical"])
    path = tmp_path / "sample.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def xlsx_with_blanks(tmp_path: Path) -> Path:
    """XLSX with blank rows interspersed between data rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Ticket Report"])
    ws.append(["Ticket number", "Ticket description", "Status", "Current Owner", "Issue type"])
    ws.append(["TSR-001", "Login issue", "Open", "PACS team", "Blocker"])
    ws.append([None, None, None, None, None])
    ws.append(["TSR-002", "Crash", "Resolved", "Product", "Critical"])
    path = tmp_path / "blanks.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def xlsx_no_ticket_col(tmp_path: Path) -> Path:
    """XLSX missing the required 'Ticket number' column."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Report"])
    ws.append(["ID", "Description", "Status"])
    ws.append([1, "Something", "Open"])
    path = tmp_path / "no_col.xlsx"
    wb.save(path)
    return path
