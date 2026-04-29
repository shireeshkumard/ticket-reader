import openpyxl
import sys

def read_tickets(filepath="Tickets.xlsx"):
    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        # Headers are on row 2; find the "Ticket number" column
        headers = [cell.value for cell in ws[2]]
        ticket_col_idx = next(
            (i for i, h in enumerate(headers) if h and "ticket number" in h.lower()),
            None
        )

        if ticket_col_idx is None:
            print(f"Could not find 'Ticket number' column. Headers found: {headers}")
            sys.exit(1)

        header_value = headers[ticket_col_idx]
        print("Ticket Numbers:")
        for row in ws.iter_rows(min_row=3, values_only=True):
            ticket = row[ticket_col_idx]
            if ticket and ticket != header_value:
                print(ticket)

    except FileNotFoundError:
        print(f"Error: '{filepath}' not found.")
        sys.exit(1)

if __name__ == "__main__":
    filepath = sys.argv[1] if len(sys.argv) > 1 else "Tickets.xlsx"
    read_tickets(filepath)
